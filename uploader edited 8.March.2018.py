#get the path for my document 
def get_dir():
    import ctypes.wintypes
    buf= ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
    ctypes.windll.shell32.SHGetFolderPathW(None, 5, None, 0, buf)
    return buf.value

#check if ZD FILES folder is available in my document
#and create it if it doesn't exist
def file_to_dir():
    import os
    d=get_dir()
    try:
        os.chdir(d + '\\ZD FILES')
        print('\nDirectory found:' + d + '\\ZD FILES')
    except FileNotFoundError:
        print('directory not found, CREATING.......')
        os.mkdir(d + '\\ZD FILES')
        os.chdir(d + '\\ZD FILES')
        print('Done! Directory successfully created!')
    return (d + '\\ZD FILES')

def get_request():
    import os, json, requests
    api_token = ..........
    email = .........
    user = email + '/token'
    creds = (user, api_token)
    url = 'https://oneacrefund-rw.zendesk.com/api/v2/'
    url_ext = 'ticket_fields.json'
    try:
        req = requests.get(url + url_ext, auth = creds)
        i = {}
        for x in req.json()['ticket_fields']:
            i[x['title']] = x['id']
        return i
    except:
        print('\nFAILED TO CONNECT TO THE INTERNET!')
    
    
def get_table(map, ws):
    ld = []
    for row in ws.iter_rows(min_row = 1, max_row = 1):
        for cell in row:
            if cell.value in map.keys():
                ld.append(cell.value)
    table = {}
    for val in ld:
        table[map[val]] = []
    for col in ws.iter_cols():
        if col[0].value in map.keys():
            for cell in col:
                if col.index(cell) != 0:
                    table[map[col[0].value]].append(cell.value or '')
    return table
	
def get_fields(table):
    l = []
    for x in range(len(table[list(table.keys())[0]])):
        l.append([])
        for key in table.keys():
            l[x].append({'id' : key, 'value' : table[key][x]})
    return l
	
def confirm_file_location():
    import os
    print('\nThe file must be placed in: ',os.getcwd(),' folder,\n \nThe folder have been created automatically if you didnt have it. \n')
    ans = input('Have this file been placed in that folder?:').lower()
    if ans == 'no':
        print('\nSorry! place the file in : ', os.getcwd(),' first and try again')

        return False
    if ans == 'yes':
        return True

#rewrite to ensure different behavior in different error situations. consider return types
def get_workbook():
    import os
    from openpyxl import Workbook, load_workbook
    try:
        fileName=input('\nEnter Workbook Name:').lower()+'.xlsx'
        os.chdir(file_to_dir())
        wb = load_workbook(fileName)
        ws = wb.active
        return ws
    except FileNotFoundError:
        print('ERROR : FILE NOT FOUND')

def pickle_function():
    custom_fields = get_fields(get_table(get_request(), get_workbook()))
    import pickle as pk
    with open('downloaded.oaf', 'wb') as f:
        pk.dump(custom_fields, f)
    del custom_fields
    del f
    with open('downloaded.oaf', 'rb') as f:
        dt=pk.load(f)
        print(dt)
        	
def main():
    import os
    os.chdir(file_to_dir())
    file_placed = confirm_file_location()
    while True:
        if file_placed==True:          
            pickle_function()
            break
        if file_placed==False:
            check=input('Do you want to continue?:')

        if check=='yes':
            pickle_function()
        else:
            print('Thank you')
            break      
	
if __name__ == "__main__":
	main()
